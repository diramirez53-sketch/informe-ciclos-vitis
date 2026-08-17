#!/usr/bin/env python3
"""
Genera el informe HTML de ciclos de cosecha a partir del Excel.

Uso:
    python generar_informe.py

Lee   : datos/2026_Seguimiento_Ciclos_de_Cosecha.xlsx  (hoja "Registro Diario")
Genera: index.html   (listo para publicar)

El corte de la serie es automatico: llega hasta HOY o hasta la ultima
columna con datos reales, lo que ocurra primero. Asi el informe nunca
muestra ciclos inflados de fechas futuras que aun no se han cosechado.
"""

import datetime
import json
import pathlib
import sys

import openpyxl

AQUI = pathlib.Path(__file__).parent
CARPETA_DATOS = AQUI
PLANTILLA = AQUI / "plantilla.html"
SALIDA = AQUI / "index.html"

# --- Estructura de la hoja "Registro Diario" -------------------------------
FILA_ENCABEZADO = 1          # fila con las fechas
PRIMERA_FILA_LOTE = 2        # primera fila de datos
COL_PRIMERA_FECHA = 23       # columna W: primera columna de fechas
COLUMNAS = {                 # nombre_campo: numero de columna
    "lote": 5,
    "plant": 1,
    "zona": 2,
    "finca": 3,
    "has": 10,
    "mat": 6,
    "casa": 7,
    "siembra": 8,
    "edad": 9,
    "palmas": 12,
    "prop": 14,
    "estado": 16,
    "ultCorte": 17,
}
COLS_TEXTO = ["plant", "finca", "mat", "casa", "prop", "estado", "ultCorte"]
ORDEN_CAMPOS = ["lote", "plant", "zona", "finca", "has", "mat", "casa",
                "siembra", "edad", "palmas", "prop", "estado", "ultCorte"]


def buscar_excel():
    """Devuelve la ruta del .xlsx en la carpeta datos/."""
    if not CARPETA_DATOS.is_dir():
        sys.exit(f"ERROR: no existe la carpeta {CARPETA_DATOS}")
    libros = [p for p in CARPETA_DATOS.glob("*.xlsx")
              if not p.name.startswith("~$")]
    if not libros:
        sys.exit(f"ERROR: no hay ningun .xlsx dentro de {CARPETA_DATOS}")
    if len(libros) > 1:
        print(f"AVISO: hay {len(libros)} libros en datos/. Se usa el mas reciente.")
        libros.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return libros[0]


def leer_parametros(wb):
    """Lee ciclo objetivo y ventana desde la hoja Parametros."""
    objetivo, ventana = 12, 5
    if "Parametros" in wb.sheetnames:
        ws = wb["Parametros"]
        for fila in range(1, ws.max_row + 1):
            etiqueta = str(ws.cell(fila, 1).value or "").lower()
            valor = ws.cell(fila, 2).value
            if isinstance(valor, (int, float)):
                if "ciclo objetivo" in etiqueta:
                    objetivo = int(valor)
                elif "ventana" in etiqueta:
                    ventana = int(valor)
    return objetivo, ventana


def extraer(ruta_excel):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    if "Registro Diario" not in wb.sheetnames:
        sys.exit('ERROR: el libro no tiene la hoja "Registro Diario"')
    ws = wb["Registro Diario"]

    objetivo, ventana = leer_parametros(wb)

    # --- columnas de fecha -------------------------------------------------
    todas = []
    for col in range(COL_PRIMERA_FECHA, ws.max_column + 1):
        valor = ws.cell(FILA_ENCABEZADO, col).value
        if isinstance(valor, datetime.datetime):
            todas.append((col, valor))
    if not todas:
        sys.exit("ERROR: no se encontraron columnas de fecha en la fila 1")

    hoy = datetime.datetime.combine(datetime.date.today(), datetime.time())

    # ultima columna con algun dato real
    ultima_con_datos = None
    for col, fecha in todas:
        if any(ws.cell(f, col).value not in (None, "")
               for f in range(PRIMERA_FILA_LOTE, ws.max_row + 1)):
            ultima_con_datos = fecha

    corte = min(hoy, ultima_con_datos) if ultima_con_datos else hoy
    cols_fecha = [(c, f) for c, f in todas if f <= corte]
    if not cols_fecha:
        sys.exit("ERROR: no hay columnas de fecha con datos hasta hoy")
    fechas = [f.strftime("%Y-%m-%d") for _, f in cols_fecha]

    # --- lotes y matriz ----------------------------------------------------
    lotes, matriz = [], []
    for fila in range(PRIMERA_FILA_LOTE, ws.max_row + 1):
        # Una fila es un lote solo si tiene nombre, zona numerica y hectareas.
        # Asi se descartan los bloques de resumen que van debajo de la tabla.
        nombre = ws.cell(fila, COLUMNAS["lote"]).value
        zona = ws.cell(fila, COLUMNAS["zona"]).value
        hectareas = ws.cell(fila, COLUMNAS["has"]).value
        if not nombre or not isinstance(zona, (int, float)) \
                or not isinstance(hectareas, (int, float)):
            continue
        registro = {}
        for campo, col in COLUMNAS.items():
            valor = ws.cell(fila, col).value
            if campo == "ultCorte":
                valor = valor.strftime("%Y-%m-%d") if isinstance(
                    valor, datetime.datetime) else None
            elif campo == "has":
                valor = round(valor or 0, 2)
            registro[campo] = valor
        lotes.append(registro)
        matriz.append([
            (v if isinstance(v, (int, float)) else None)
            for v in (ws.cell(fila, c).value for c, _ in cols_fecha)
        ])

    if not lotes:
        sys.exit("ERROR: no se encontro ningun lote con nombre")

    # --- compactar textos repetidos (reduce el peso del HTML) --------------
    tablas = {}
    for campo in COLS_TEXTO:
        vistos = []
        for lote in lotes:
            v = lote.get(campo)
            if v is not None and v not in vistos:
                vistos.append(v)
        tablas[campo] = vistos

    filas = []
    for lote in lotes:
        fila_out = []
        for campo in ORDEN_CAMPOS:
            v = lote.get(campo)
            if campo in COLS_TEXTO:
                fila_out.append(tablas[campo].index(v) if v is not None else -1)
            else:
                fila_out.append(v)
        filas.append(fila_out)

    return {
        "dates": fechas,
        "K": ORDEN_CAMPOS,
        "T": tablas,
        "L": filas,
        "matrix": matriz,
        "cicloObjetivo": objetivo,
        "ventana": ventana,
        "corte": fechas[-1],
    }, lotes, matriz


def main():
    ruta_excel = buscar_excel()
    print(f"Leyendo {ruta_excel.name} ...")
    datos, lotes, matriz = extraer(ruta_excel)

    if not PLANTILLA.exists():
        sys.exit(f"ERROR: falta {PLANTILLA.name}")
    html = PLANTILLA.read_text(encoding="utf-8")
    if "/*__DATA__*/" not in html:
        sys.exit("ERROR: la plantilla no contiene el marcador /*__DATA__*/")

    html = html.replace(
        "/*__DATA__*/",
        json.dumps(datos, separators=(",", ":"), ensure_ascii=False))
    SALIDA.write_text(html, encoding="utf-8")

    # --- resumen de control ------------------------------------------------
    ultimo = len(datos["dates"]) - 1
    obj = datos["cicloObjetivo"]
    con_dato = [(l, f[ultimo]) for l, f in zip(lotes, matriz)
                if f[ultimo] is not None]
    has = sum(l["has"] for l, _ in con_dato)
    fuera = sum(l["has"] for l, v in con_dato if v > obj)
    pond = (sum(l["has"] * v for l, v in con_dato) / has) if has else 0

    print(f"OK  {SALIDA.name}  ({SALIDA.stat().st_size / 1024:.1f} KB)")
    print(f"    Serie      : {datos['dates'][0]} -> {datos['dates'][-1]} "
          f"({len(datos['dates'])} dias)")
    print(f"    Lotes      : {len(lotes)} ({len(con_dato)} con ciclo)")
    print(f"    Has        : {has:,.1f} con seguimiento")
    print(f"    Fuera ciclo: {fuera:,.1f} has  ({fuera / has * 100:.1f}%)"
          if has else "")
    print(f"    Ponderado  : {pond:.1f} d  (objetivo {obj} d)")

    # --- control de frescura ----------------------------------------------
    # Un corte real se reconoce cuando el contador de un lote vuelve a 0.
    # Si hace varios dias que no hay ninguno, el libro dejo de alimentarse:
    # los ciclos siguen subiendo solos y el informe se ve peor de lo real.
    ultimo_corte = None
    for j in range(len(datos["dates"]) - 1, -1, -1):
        if any(fila[j] == 0 for fila in matriz):
            ultimo_corte = datos["dates"][j]
            break

    if ultimo_corte:
        dias = (datetime.date.fromisoformat(datos["dates"][-1])
                - datetime.date.fromisoformat(ultimo_corte)).days
        print(f"    Ult. corte : {ultimo_corte}  ({dias} dias atras)")
        if dias >= 3:
            print()
            print("  " + "!" * 60)
            print(f"  AVISO: hace {dias} dias que no se registra ningun corte.")
            print("  Verifique que el Excel este actualizado: si no lo esta,")
            print("  los ciclos suben solos y el informe exagera el atraso.")
            print("  " + "!" * 60)
    else:
        print("    AVISO: no se detecto ningun corte en toda la serie.")


if __name__ == "__main__":
    main()
